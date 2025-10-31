import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import logging
from datetime import datetime
from utils import get_colombian_datetime_now

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Get the directory where this script is located
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# Get the parent directory (Viaticos folder)
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
# Change to project directory if needed for relative paths
# os.chdir(PROJECT_DIR)  # Uncomment if needed for relative path operations

def consolidate_excel_sheets(input_file, output_file=None):
    """
    Consolida todas las hojas de un archivo Excel en una sola hoja con formato profesional.

    Args:
        input_file (str): Ruta al archivo Excel de entrada
        output_file (str): Ruta al archivo Excel de salida (opcional)

    Returns:
        str: Ruta del archivo consolidado creado
    """

    logger.info("🔄 Iniciando consolidación de archivo Excel...")
    logger.info(f"📂 Archivo de entrada: {input_file}")

    # Verificar que el archivo existe
    if not os.path.exists(input_file):
        logger.error(f"El archivo {input_file} no existe")
        raise FileNotFoundError(f"El archivo {input_file} no existe")

    # Crear nombre de archivo de salida si no se proporciona
    if output_file is None:
        base_name = os.path.splitext(input_file)[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"{base_name}_consolidado_{timestamp}.xlsx"

    try:
        # Leer el archivo Excel y obtener todas las hojas
        excel_file = pd.ExcelFile(input_file)
        all_sheets = excel_file.sheet_names

        logger.info(f"📊 Hojas encontradas: {len(all_sheets)}")
        logger.info(f"📋 Nombres de hojas: {all_sheets}")

        # Filtrar hojas que no son de datos (excluir LIST y hojas similares)
        data_sheets = [sheet for sheet in all_sheets if sheet.upper() not in ['LIST', 'SUMMARY', 'RESUMEN', 'INDEX']]

        logger.info(f"📈 Hojas de datos a procesar: {len(data_sheets)}")

        # Lista para almacenar todos los datos consolidados
        consolidated_data = []
        headers_set = False
        headers = None

        # Procesar cada hoja
        for i, sheet_name in enumerate(data_sheets):
            logger.info(f"🔄 Procesando hoja {i + 1}/{len(data_sheets)}: {sheet_name}")

            try:
                # Leer la hoja actual
                df = pd.read_excel(input_file, sheet_name=sheet_name)

                # Buscar la fila de encabezados (usualmente contiene "IDENTIFICACIÓN" o "No DE IDENTIFICACIÓN")
                header_row = None
                for idx, row in df.iterrows():
                    if any('IDENTIFICACIÓN' in str(cell).upper() for cell in row if pd.notna(cell)):
                        header_row = idx
                        break

                if header_row is None:
                    logger.warning(f"⚠️  No se encontraron encabezados válidos en {sheet_name}, saltando...")
                    continue

                # Releer la hoja usando la fila de encabezados correcta
                df = pd.read_excel(input_file, sheet_name=sheet_name, header=header_row)

                # Limpiar el DataFrame
                # Eliminar filas completamente vacías
                df = df.dropna(how='all')

                # Eliminar filas que no tienen datos esenciales (sin ID de identificación)
                id_columns = [col for col in df.columns if 'IDENTIFICACIÓN' in str(col).upper()]
                if id_columns:
                    df = df.dropna(subset=id_columns, how='all')

                if df.empty:
                    logger.warning(f"⚠️  No hay datos válidos en {sheet_name}, saltando...")
                    continue

                # Agregar columna SEDE al inicio
                df.insert(0, 'SEDE', sheet_name.strip())

                # Establecer encabezados en la primera iteración
                if not headers_set:
                    headers = list(df.columns)
                    headers_set = True
                    logger.info(f"📋 Encabezados establecidos: {len(headers)} columnas")

                # Asegurar que todas las hojas tengan las mismas columnas
                current_headers = list(df.columns)
                if current_headers != headers:
                    logger.warning(f"⚠️  Ajustando columnas para {sheet_name}")
                    # Reindexar para que coincidan las columnas
                    df = df.reindex(columns=headers)

                # Agregar los datos al consolidado
                consolidated_data.append(df)
                logger.info(f"✅ {sheet_name}: {len(df)} registros agregados")

            except Exception as e:
                logger.error(f"❌ Error procesando {sheet_name}: {str(e)}")
                continue

        # Verificar que tenemos datos para consolidar
        if not consolidated_data:
            raise ValueError("No se encontraron datos válidos para consolidar")

        # Concatenar todos los DataFrames
        logger.info("🔄 Consolidando todos los datos...")
        final_df = pd.concat(consolidated_data, ignore_index=True)

        # Limpiar y formatear datos finales
        logger.info("🧹 Limpiando datos finales...")

        # Rellenar valores NaN con cadenas vacías para mejor presentación
        final_df = final_df.fillna('')

        # Convertir columnas numéricas apropiadamente
        numeric_columns = ['REC', 'VIGENCIA', 'No. ORDEN', 'DIAS OC', 'VALOR VIÁTICO DIARIO',
                           'VALOR VIATICOS ORDEN', 'GASTOS ORDEN', 'TOTAL GIRADO ORDEN',
                           'DIAS LEG.', 'VIATICOS LEG.', 'GASTOS LEG.', 'TOTAL LEG.', 'TOTAL REINTEGRO']

        for col in numeric_columns:
            if col in final_df.columns:
                try:
                    final_df[col] = pd.to_numeric(final_df[col], errors='coerce').fillna(0)
                except (ValueError, TypeError):
                    continue

        # Crear archivo Excel con formato profesional
        logger.info("📝 Creando archivo Excel con formato profesional...")

        # Crear workbook y worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos Consolidados"

        # Agregar datos al worksheet
        for r in dataframe_to_rows(final_df, index=False, header=True):
            ws.append(r)

        # Aplicar formato profesional
        apply_professional_formatting(ws, final_df)

        # Crear hoja de resumen
        create_summary_sheet(wb, final_df, data_sheets)

        # Guardar archivo
        wb.save(output_file)

        # Mostrar estadísticas finales
        logger.info("\n" + "=" * 60)
        logger.info("✅ CONSOLIDACIÓN COMPLETADA EXITOSAMENTE")
        logger.info("=" * 60)
        logger.info(f"📁 Archivo creado: {output_file}")
        logger.info(f"📊 Total de hojas procesadas: {len(data_sheets)}")
        logger.info(f"📈 Total de registros: {len(final_df):,}")
        logger.info(f"📋 Total de columnas: {len(final_df.columns)}")
        logger.info(f"💾 Tamaño del archivo: {os.path.getsize(output_file) / (1024 * 1024):.2f} MB")

        # Mostrar estadísticas por sede
        logger.info("\n📊 REGISTROS POR SEDE:")
        logger.info("-" * 40)
        sede_counts = final_df['SEDE'].value_counts().sort_values(ascending=False)
        for sede, count in sede_counts.items():
            logger.info(f"  {sede}: {count:,} registros")

        logger.info(f"\n🎯 Archivo Excel consolidado creado exitosamente: {output_file}")

        return output_file

    except Exception as e:
        logger.error(f"❌ Error durante la consolidación: {str(e)}")
        raise

def apply_professional_formatting(ws, df):
    """Aplica formato profesional al worksheet"""

    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(size=10)
    data_alignment = Alignment(horizontal="left", vertical="center")

    # Bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Formatear encabezados (fila 1)
    for col_num, column in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Formatear datos
    for row_num in range(2, len(df) + 2):
        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

            # Alternar colores de filas
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
        ws.column_dimensions[column_letter].width = adjusted_width

    # Congelar paneles (primera fila)
    ws.freeze_panes = 'A2'

    # Agregar filtros automáticos
    ws.auto_filter.ref = ws.dimensions

def create_summary_sheet(wb, df, sheet_names):
    """Crea una hoja de resumen con estadísticas"""

    # Crear nueva hoja
    summary_ws = wb.create_sheet(title="Resumen")

    # Título
    summary_ws['A1'] = "RESUMEN DE CONSOLIDACIÓN"
    summary_ws['A1'].font = Font(bold=True, size=16, color="366092")

    # Información general
    row = 3
    summary_ws[f'A{row}'] = "Información General"
    summary_ws[f'A{row}'].font = Font(bold=True, size=12)

    row += 1
    summary_ws[f'A{row}'] = "Fecha de consolidación:"
    summary_ws[f'B{row}'] = get_colombian_datetime_now()

    row += 1
    summary_ws[f'A{row}'] = "Total de registros:"
    summary_ws[f'B{row}'] = len(df)

    row += 1
    summary_ws[f'A{row}'] = "Total de columnas:"
    summary_ws[f'B{row}'] = len(df.columns)

    row += 1
    summary_ws[f'A{row}'] = "Total de sedes:"
    summary_ws[f'B{row}'] = len(sheet_names)

    # Estadísticas por sede
    row += 3
    summary_ws[f'A{row}'] = "Registros por Sede"
    summary_ws[f'A{row}'].font = Font(bold=True, size=12)

    row += 1
    summary_ws[f'A{row}'] = "Sede"
    summary_ws[f'B{row}'] = "Cantidad"
    summary_ws[f'A{row}'].font = Font(bold=True)
    summary_ws[f'B{row}'].font = Font(bold=True)

    sede_counts = df['SEDE'].value_counts().sort_values(ascending=False)
    for sede, count in sede_counts.items():
        row += 1
        summary_ws[f'A{row}'] = sede
        summary_ws[f'B{row}'] = count

    # Ajustar anchos de columna
    summary_ws.column_dimensions['A'].width = 30
    summary_ws.column_dimensions['B'].width = 15


# Función principal para ejecutar
def main():
    """Función principal para ejecutar la consolidación"""

    # Configurar el nombre del archivo usando rutas relativas al proyecto
    input_file = os.path.join(PROJECT_DIR, "Data", "02_REPORTE VIATICOS DT 2025.xlsx")

    logger.info("🚀 CONSOLIDADOR DE EXCEL - REPORTE VIÁTICOS DT 2025")
    logger.info("=" * 60)

    try:
        # Verificar que el archivo existe
        if not os.path.exists(input_file):
            logger.error(f"❌ Error: El archivo '{input_file}' no se encuentra.")
            logger.error(f"📁 Verifica que exista la carpeta Data en: {PROJECT_DIR}")
            return

        # Ejecutar consolidación
        output_file = consolidate_excel_sheets(input_file)

        logger.info(f"\n🎉 ¡Consolidación completada exitosamente!")
        logger.info(f"📁 Archivo creado: {output_file}")

    except Exception as e:
        logger.error(f"❌ Error durante la ejecución: {str(e)}")
        raise


# Ejecutar si se corre directamente
if __name__ == "__main__":
    main()